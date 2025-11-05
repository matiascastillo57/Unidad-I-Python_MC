# monitoring/export_views.py
"""
<<<<<<< HEAD
Vistas para exportar datos a Excel
"""
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Zone, Device, Category, Measurement
from ecoenergy.decorators import permission_required_with_message


def get_user_organization(user):
    """Helper para obtener organización del usuario"""
=======
Vistas para exportar datos a Excel (.xlsx)
"""
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from .models import Device, Zone, Category, Measurement


def get_user_organization(user):
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
    if user.is_superuser:
        return None
    try:
        return user.userprofile.organization
    except:
        return None


<<<<<<< HEAD
def apply_excel_styles(ws):
    """Aplicar estilos profesionales al Excel"""
    # Estilo para encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Aplicar a la primera fila
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar bordes a todas las celdas con datos
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


@login_required
@permission_required_with_message('monitoring.view_zone')
def export_zones_excel(request):
    """Exportar zonas a Excel"""
    organization = get_user_organization(request.user)
    
    # Obtener datos
    if organization:
        zones = Zone.objects.filter(organization=organization, state='ACTIVE').select_related('organization')
    else:
        zones = Zone.objects.filter(state='ACTIVE').select_related('organization')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Zonas"
    
    # Headers
    headers = ['ID', 'Nombre', 'Descripción', 'Organización', 'Fecha Creación', 'Estado']
    ws.append(headers)
    
    # Data
    for zone in zones:
        ws.append([
            zone.id,
            zone.name,
            zone.description or '',
            zone.organization.name,
            zone.created_at.strftime('%d/%m/%Y %H:%M'),
            'Activa' if zone.state == 'ACTIVE' else 'Inactiva'
        ])
    
    # Aplicar estilos
    apply_excel_styles(ws)
    
    # Crear response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'zonas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


@login_required
@permission_required_with_message('monitoring.view_device')
=======
@login_required
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
def export_devices_excel(request):
    """Exportar dispositivos a Excel"""
    organization = get_user_organization(request.user)
    
<<<<<<< HEAD
    # Obtener datos
    if organization:
        devices = Device.objects.filter(
            organization=organization, 
            state='ACTIVE'
        ).select_related('category', 'zone', 'organization')
    else:
        devices = Device.objects.filter(state='ACTIVE').select_related('category', 'zone', 'organization')
=======
    # Obtener dispositivos
    if organization:
        devices = Device.objects.filter(organization=organization, state='ACTIVE')
    else:
        devices = Device.objects.filter(state='ACTIVE')
    
    devices = devices.select_related('category', 'zone').order_by('name')
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispositivos"
    
<<<<<<< HEAD
    # Headers
    headers = ['ID', 'Nombre', 'Categoría', 'Zona', 'Consumo Máximo (kW)', 'Organización', 'Fecha Creación', 'Estado']
    ws.append(headers)
    
    # Data
=======
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Encabezados
    headers = ['ID', 'Nombre', 'Categoría', 'Zona', 'Consumo Máx (kW)', 'Estado', 'Creado']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
    for device in devices:
        ws.append([
            device.id,
            device.name,
            device.category.name,
            device.zone.name,
            float(device.max_consumption),
<<<<<<< HEAD
            device.organization.name,
            device.created_at.strftime('%d/%m/%Y %H:%M'),
            'Activo' if device.state == 'ACTIVE' else 'Inactivo'
        ])
    
    # Aplicar estilos
    apply_excel_styles(ws)
    
    # Crear response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'dispositivos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
=======
            device.get_state_display(),
            device.created_at.strftime('%d/%m/%Y %H:%M')
        ])
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'dispositivos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_zones_excel(request):
    """Exportar zonas a Excel"""
    organization = get_user_organization(request.user)
    
    if organization:
        zones = Zone.objects.filter(organization=organization, state='ACTIVE')
    else:
        zones = Zone.objects.filter(state='ACTIVE')
    
    zones = zones.order_by('name')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Zonas"
    
    # Estilos
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    headers = ['ID', 'Nombre', 'Descripción', 'Organización', 'Estado', 'Creado']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for zone in zones:
        ws.append([
            zone.id,
            zone.name,
            zone.description or '-',
            zone.organization.name,
            zone.get_state_display(),
            zone.created_at.strftime('%d/%m/%Y %H:%M')
        ])
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'zonas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
    
    wb.save(response)
    return response


@login_required
<<<<<<< HEAD
@permission_required_with_message('monitoring.view_category')
=======
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
def export_categories_excel(request):
    """Exportar categorías a Excel"""
    organization = get_user_organization(request.user)
    
<<<<<<< HEAD
    # Obtener datos
    if organization:
        categories = Category.objects.filter(organization=organization, state='ACTIVE').select_related('organization')
    else:
        categories = Category.objects.filter(state='ACTIVE').select_related('organization')
    
    # Crear workbook
=======
    if organization:
        categories = Category.objects.filter(organization=organization, state='ACTIVE')
    else:
        categories = Category.objects.filter(state='ACTIVE')
    
    categories = categories.order_by('name')
    
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías"
    
<<<<<<< HEAD
    # Headers
    headers = ['ID', 'Nombre', 'Descripción', 'Organización', 'Fecha Creación', 'Estado']
    ws.append(headers)
    
    # Data
    for category in categories:
        ws.append([
            category.id,
            category.name,
            category.description or '',
            category.organization.name,
            category.created_at.strftime('%d/%m/%Y %H:%M'),
            'Activa' if category.state == 'ACTIVE' else 'Inactiva'
        ])
    
    # Aplicar estilos
    apply_excel_styles(ws)
    
    # Crear response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'categorias_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
=======
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=12)
    
    headers = ['ID', 'Nombre', 'Descripción', 'Organización', 'Estado', 'Creado']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for cat in categories:
        ws.append([
            cat.id,
            cat.name,
            cat.description or '-',
            cat.organization.name,
            cat.get_state_display(),
            cat.created_at.strftime('%d/%m/%Y %H:%M')
        ])
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'categorias_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
    
    wb.save(response)
    return response


@login_required
<<<<<<< HEAD
@permission_required_with_message('monitoring.view_measurement')
=======
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
def export_measurements_excel(request):
    """Exportar mediciones a Excel"""
    organization = get_user_organization(request.user)
    
<<<<<<< HEAD
    # Obtener datos
    if organization:
        measurements = Measurement.objects.filter(
            organization=organization, 
            state='ACTIVE'
        ).select_related('device', 'device__zone', 'device__category').order_by('-measurement_date')[:1000]
    else:
        measurements = Measurement.objects.filter(state='ACTIVE').select_related(
            'device', 'device__zone', 'device__category'
        ).order_by('-measurement_date')[:1000]
    
    # Crear workbook
=======
    if organization:
        measurements = Measurement.objects.filter(organization=organization, state='ACTIVE')
    else:
        measurements = Measurement.objects.filter(state='ACTIVE')
    
    measurements = measurements.select_related('device', 'device__zone').order_by('-measurement_date')[:500]
    
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
    wb = Workbook()
    ws = wb.active
    ws.title = "Mediciones"
    
<<<<<<< HEAD
    # Headers
    headers = [
        'ID', 'Dispositivo', 'Categoría', 'Zona', 'Consumo (kW)', 
        'Consumo Máximo (kW)', '¿Excede?', 'Fecha Medición', 'Notas'
    ]
    ws.append(headers)
    
    # Data
    for m in measurements:
        excede = 'SÍ' if m.consumption_value > m.device.max_consumption else 'NO'
        ws.append([
            m.id,
            m.device.name,
            m.device.category.name,
            m.device.zone.name,
            float(m.consumption_value),
            float(m.device.max_consumption),
            excede,
            m.measurement_date.strftime('%d/%m/%Y %H:%M:%S'),
            m.notes or ''
        ])
    
    # Aplicar estilos
    apply_excel_styles(ws)
    
    # Colorear filas que exceden límite
    from openpyxl.styles import PatternFill
    red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2):  # Saltar header
        if row[6].value == 'SÍ':  # Columna "¿Excede?"
            for cell in row:
                cell.fill = red_fill
    
    # Crear response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'mediciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
=======
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    headers = ['ID', 'Dispositivo', 'Zona', 'Consumo (kW)', 'Fecha/Hora', 'Notas', 'Estado']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for m in measurements:
        ws.append([
            m.id,
            m.device.name,
            m.device.zone.name,
            float(m.consumption_value),
            m.measurement_date.strftime('%d/%m/%Y %H:%M:%S'),
            m.notes or '-',
            m.get_state_display()
        ])
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 12
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'mediciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
>>>>>>> ff1242244faefc3c9a65023a6e41b49b1ca4453e
    
    wb.save(response)
    return response